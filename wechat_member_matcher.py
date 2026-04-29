#!/usr/bin/env python3
"""
WeChat Group Member → Member Database Matcher
=============================================
Matches WeChat group member screenshots (manually transcribed to a list)
against a member CSV database, producing a 7-tab Excel report.

USAGE:
    python3 wechat_member_matcher.py \
        --csv   all_members_2026-04-25.csv \
        --wechat wechat_members.csv \
        --output wechat_matching_output.xlsx

INPUT FILES:
    1. Member CSV (--csv):
       Standard export with columns:
         District, Member ID, Full Name, Status, Expiration, Gender,
         WeChat ID, Email, Type, Family ID, Payment Date,
         Membership Fee Paid, Payment Transaction

       Key matching fields:
         - Member ID: format A0001 – A9999
         - WeChat ID: member's self-reported WeChat username
         - Full Name: English name
         - Email: may contain name fragments in local part

    2. WeChat members CSV (--wechat):
       Two-column CSV transcribed from WeChat group member screenshots:
         wechat_name, wechat_alias
       
       "wechat_name"  = the large display name shown in WeChat
       "wechat_alias" = the nickname shown below (昵称 or 群昵称)
       Leave alias blank if not shown.

       Example rows:
         贾森（Zhaoxun Liu）,贾森
         A0019 Jimmy 🔵,Jimmy 🔵
         珊妹子,
         龙在纽约 付龙昌,龙昌 A0014

OUTPUT: 7-tab Excel workbook
    Tab 1: Confirmed Active    – confirmed match + active/lifetime status
    Tab 2: Confirmed Review    – confirmed match but inactive/expired (still in group!)
    Tab 3: Guessed Active      – probable match + active/lifetime, needs human review
    Tab 4: Guessed Review      – probable match + inactive/expired, needs human review
    Tab 5: WeChat No CSV Match – in WeChat group but no CSV record found
    Tab 6: CSV No WeChat Active– active members not found in WeChat group
    Tab 7: CSV No WeChat Rest  – inactive/expired members not in WeChat group

MATCHING LOGIC (in priority order):
    1. Member ID in name/alias  : "A0121" or "A 0121" or "timA0293" → direct lookup
    2. WeChat ID exact match    : CSV WeChat ID == alias or display name (any length)
    3. WeChat ID substring      : CSV WeChat ID (≥4 chars) contained in alias (≥4 chars)
    4. Full name in WeChat      : CSV full name (≥4 chars) found in display or alias
    5. WeChat name in full name : WC display (≥4 chars) found in CSV full name
    6. WeChat alias in full name: WC alias (≥4 chars) found in CSV full name
    7. Name part match          : individual name tokens (≥4 chars) found in WC display/alias
    8. Email match              : 2+ tokens match email local part, OR 1 unique token ≥6 chars

    Scores: 95=confirmed, 85/80/75/72/68/65/60=guessed (descending confidence)
    Each CSV member ID appears at most once (highest confidence kept).
"""

import csv
import re
import sys
import argparse
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


# ─────────────────────────────────────────────────────────────────────────────
# CORE MATCHING FUNCTIONS
# ─────────────────────────────────────────────────────────────────────────────

def extract_member_id(s):
    """
    Extract member ID (A0001–A9999) from a string.
    Handles: 'A0121', 'A 0121' (space), 'timA0293' (no boundary), 'A383' (3-digit).
    """
    if not s:
        return None
    m = re.search(r'A\s*0*(\d{3,4})(?:\b|\D|$)', s, re.IGNORECASE)
    if m:
        return f"A{int(m.group(1)):04d}"
    return None


def normalize(s):
    """Strip punctuation/spaces and lowercase for fuzzy comparison."""
    if not s:
        return ''
    return re.sub(r'[\s\-_\*\(\)\[\]\.,，。·•]+', '', s).lower()


def is_active(member):
    """True if member status is active or lifetime."""
    return member.get('Status', '').lower().strip() in ('active', 'lifetime')


def email_match_score(wc_name, wc_alias, email):
    """
    Score a match based on name tokens appearing in the email local part.
    Requires 2+ tokens OR 1 unique long token (≥6 chars) to avoid false positives.
    Common short tokens like 'zhang', 'wang' alone are NOT enough.
    Returns (score, reason_string).
    """
    if not email or '@' not in email:
        return 0, ''
    local = email.split('@')[0].lower()
    all_text = (wc_name or '') + ' ' + (wc_alias or '')
    # Only English tokens ≥4 chars
    tokens = [t.lower() for t in re.findall(r'[a-zA-Z]{4,}', all_text)]
    if not tokens:
        return 0, ''
    matched = [t for t in tokens if t in local]
    if len(matched) >= 2:
        return 72, f"Email match: {' + '.join(matched)} in {email}"
    if len(matched) == 1 and len(matched[0]) >= 6:
        return 68, f"Email match: unique token '{matched[0]}' in {email}"
    if len(matched) == 1 and len(tokens) == 1 and len(matched[0]) >= 5:
        return 65, f"Email match: '{matched[0]}' in {email}"
    return 0, ''


MIN_LEN = 4  # minimum chars for substring fuzzy matching


def score_member_match(wc_name, wc_alias, member):
    """
    Score how well a CSV member record matches a WeChat entry.
    Returns (score 0–95, reason_string).
    """
    wc_norm       = normalize(wc_name)
    wc_alias_norm = normalize(wc_alias) if wc_alias else ''
    full_name     = member['Full Name'].strip()
    wechat_id     = member['WeChat ID'].strip()
    email         = member['Email'].strip()
    fn_norm       = normalize(full_name)
    wid_norm      = normalize(wechat_id)

    # Rule A: CSV WeChat ID exact-matches alias (any non-empty length)
    if wechat_id and wc_alias and wid_norm == wc_alias_norm:
        return 95, f"CSV WeChat ID '{wechat_id}' = alias (exact)"

    # Rule B: CSV WeChat ID exact-matches display name (any non-empty length)
    if wechat_id and wc_norm and wid_norm == wc_norm:
        return 95, f"CSV WeChat ID '{wechat_id}' = display name (exact)"

    # Rule C: CSV WeChat ID (≥4 chars) contained in alias (≥4 chars)
    if (wechat_id and wc_alias
            and len(wid_norm) >= MIN_LEN and len(wc_alias_norm) >= MIN_LEN
            and wid_norm in wc_alias_norm):
        return 85, f"CSV WeChat ID '{wechat_id}' contained in alias"

    # Rule D: full name (≥4 chars) found in WC display or alias
    if (fn_norm and len(fn_norm) >= MIN_LEN
            and (fn_norm in wc_norm or fn_norm in wc_alias_norm)):
        return 80, f"Full name '{full_name}' in WeChat display"

    # Rule E: WC name (≥4 chars) found in full name
    if fn_norm and len(fn_norm) >= MIN_LEN and len(wc_norm) >= MIN_LEN and wc_norm in fn_norm:
        return 75, f"WeChat name in CSV full name '{full_name}'"

    # Rule F: WC alias (≥4 chars) found in full name
    if fn_norm and len(fn_norm) >= MIN_LEN and len(wc_alias_norm) >= MIN_LEN and wc_alias_norm in fn_norm:
        return 75, f"WeChat alias in CSV full name '{full_name}'"

    # Rule G: individual name tokens (≥4 chars) in WC display or alias
    best_g = 0
    best_r = ''
    for part in full_name.split():
        pn = normalize(part)
        if len(pn) >= MIN_LEN:
            if (len(wc_norm) >= MIN_LEN and pn in wc_norm) or \
               (len(wc_alias_norm) >= MIN_LEN and pn in wc_alias_norm):
                best_g, best_r = max(best_g, 60), f"Name part '{part}' in WeChat display"
    if best_g:
        return best_g, best_r

    # Rule H: email local part (last resort, only if score still 0)
    es, er = email_match_score(wc_name, wc_alias, email)
    if es:
        return es, er

    return 0, ''


# ─────────────────────────────────────────────────────────────────────────────
# MAIN MATCHING PIPELINE
# ─────────────────────────────────────────────────────────────────────────────

def load_csv_members(filepath):
    """Load member CSV, deduplicate by Member ID, return dict and list."""
    rows = []
    with open(filepath, encoding='utf-8-sig') as f:
        for row in csv.DictReader(f):
            rows.append(row)
    id_to_member = {}
    for row in rows:
        mid = row.get('Member ID', '').strip()
        if mid and mid not in id_to_member:
            id_to_member[mid] = row
    return id_to_member, list(id_to_member.values())


def load_wechat_members(filepath):
    """
    Load WeChat members CSV with columns: wechat_name, wechat_alias
    Deduplicate by (name, alias) pair.
    """
    entries = []
    seen = set()
    with open(filepath, encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            name  = (row.get('wechat_name') or '').strip()
            alias = (row.get('wechat_alias') or '').strip() or None
            key   = (name, alias or '')
            if name and key not in seen:
                seen.add(key)
                entries.append((name, alias))
    return entries


def run_matching(wechat_list, id_to_member, all_members):
    """
    Run full matching pipeline.
    Returns: confirmed_by_id, guessed_by_mid, wc_no_match
    """
    confirmed_by_id    = {}   # member_id -> (wc_name, wc_alias, member, reason)
    guessed_candidates = []   # list of (wc_name, wc_alias, mid, score, reason)
    wc_no_match        = []   # list of (wc_name, wc_alias)

    for wc_name, wc_alias in wechat_list:
        # Step 1: direct member ID extraction (highest priority)
        found_id = extract_member_id(wc_name) or extract_member_id(wc_alias)
        if found_id and found_id in id_to_member:
            if found_id not in confirmed_by_id:
                confirmed_by_id[found_id] = (
                    wc_name, wc_alias, id_to_member[found_id],
                    f"Member ID {found_id} in WeChat name/alias"
                )
            continue  # matched; move on

        # Step 2: fuzzy matching against all members
        best_mid, best_score, best_reason = None, 0, ''
        for member in all_members:
            mid = member['Member ID'].strip()
            score, reason = score_member_match(wc_name, wc_alias, member)
            if score > best_score:
                best_score, best_mid, best_reason = score, mid, reason

        if best_score >= 95 and best_mid and best_mid not in confirmed_by_id:
            confirmed_by_id[best_mid] = (
                wc_name, wc_alias, id_to_member[best_mid], best_reason
            )
        elif best_score >= 60 and best_mid:
            guessed_candidates.append((wc_name, wc_alias, best_mid, best_score, best_reason))
        else:
            wc_no_match.append((wc_name, wc_alias))

    # Deduplicate guessed: one entry per member ID, keep highest confidence
    guessed_by_mid = {}
    for wc_name, wc_alias, mid, conf, reason in guessed_candidates:
        if mid not in confirmed_by_id:
            if mid not in guessed_by_mid or conf > guessed_by_mid[mid][3]:
                guessed_by_mid[mid] = (wc_name, wc_alias, id_to_member[mid], conf, reason)

    return confirmed_by_id, guessed_by_mid, wc_no_match


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL OUTPUT
# ─────────────────────────────────────────────────────────────────────────────

# Column definitions
CONF_COLS  = ['WeChat Display Name', 'WeChat Alias', 'Match Reason',
              'Member ID', 'Full Name', 'Status', 'WeChat ID', 'Expiration',
              'Email', 'District', 'Gender', 'Type', 'Family ID',
              'Payment Date', 'Membership Fee Paid', 'Payment Transaction']
GUESS_COLS = ['WeChat Display Name', 'WeChat Alias', 'Confidence %', 'Guess Reason',
              'Member ID', 'Full Name', 'Status', 'WeChat ID', 'Expiration',
              'Email', 'District', 'Gender', 'Type', 'Family ID',
              'Payment Date', 'Membership Fee Paid', 'Payment Transaction']
WC_NM_COLS = ['WeChat Display Name', 'WeChat Alias', 'Note',
              'Member ID (from name)', 'Full Name', 'Status', 'District', 'Email']
CSV_COLS   = ['Member ID', 'Full Name', 'Status', 'WeChat ID', 'Expiration',
              'Email', 'District', 'Gender', 'Type', 'Family ID',
              'Payment Date', 'Membership Fee Paid', 'Payment Transaction']

CONF_W  = [30, 30, 40,  9, 22, 8, 15, 11, 28, 12,  7, 10,  9, 12, 14, 18]
GUESS_W = [30, 30, 11, 45,  9, 22, 8, 15, 11, 28, 12,  7, 10,  9, 12, 14, 18]
WC_W    = [35, 35, 42, 18, 22, 10, 14, 30]
CSV_W   = [ 9, 22,  8, 15, 11, 28, 12,  7, 10,  9, 12, 14, 18]

# Color palette
PALETTE = {
    'confirmed_active':  '1F7A4A',
    'confirmed_review':  '375623',
    'guessed_active':    'C55A11',
    'guessed_review':    'BF8F00',
    'wc_nomatch':        '2E6DAD',
    'csv_nomatch_active':'595959',
    'csv_nomatch_rest':  '808080',
}

HDR_FONT = Font(bold=True, color='FFFFFF', name='Arial', size=10)
ROW_FONT = Font(name='Arial', size=10)
ALT_FILL = PatternFill('solid', start_color='F2F2F2')
YEL_FILL = PatternFill('solid', start_color='FFF2CC')
THIN     = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'),  bottom=Side(style='thin'))


def make_fill(hex_color):
    return PatternFill('solid', start_color=hex_color)


def write_header(ws, cols, fill):
    for ci, col in enumerate(cols, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.font      = HDR_FONT
        c.fill      = fill
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border    = THIN
    ws.row_dimensions[1].height = 28


def style_data_row(ws, ri, n_cols, highlight=False):
    fill = YEL_FILL if highlight else (ALT_FILL if ri % 2 == 0 else None)
    for ci in range(1, n_cols + 1):
        c = ws.cell(row=ri, column=ci)
        c.font      = ROW_FONT
        c.border    = THIN
        c.alignment = Alignment(vertical='center')
        if fill:
            c.fill = fill


def set_widths(ws, widths):
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


def write_confirmed_sheet(ws, items, fill, highlight=False):
    write_header(ws, CONF_COLS, fill)
    for i, (wn, wa, m, reason) in enumerate(items, 2):
        ws.cell(row=i, column=1, value=wn or '')
        ws.cell(row=i, column=2, value=wa or '')
        ws.cell(row=i, column=3, value=reason or '')
        for ci, col in enumerate(CONF_COLS[3:], 4):
            ws.cell(row=i, column=ci, value=m.get(col, ''))
        style_data_row(ws, i, len(CONF_COLS), highlight)
    set_widths(ws, CONF_W)
    ws.freeze_panes = 'A2'


def write_guessed_sheet(ws, items, fill, highlight=False):
    write_header(ws, GUESS_COLS, fill)
    for i, (wn, wa, m, conf, reason) in enumerate(items, 2):
        ws.cell(row=i, column=1, value=wn or '')
        ws.cell(row=i, column=2, value=wa or '')
        ws.cell(row=i, column=3, value=conf)
        ws.cell(row=i, column=4, value=reason or '')
        for ci, col in enumerate(GUESS_COLS[4:], 5):
            ws.cell(row=i, column=ci, value=m.get(col, ''))
        style_data_row(ws, i, len(GUESS_COLS), highlight)
    set_widths(ws, GUESS_W)
    ws.freeze_panes = 'A2'


def write_wc_nomatch_sheet(ws, items, id_to_member, fill):
    write_header(ws, WC_NM_COLS, fill)
    for i, (wn, wa) in enumerate(items, 2):
        fid    = extract_member_id(wn) or extract_member_id(wa)
        member = id_to_member.get(fid) if fid else None
        ws.cell(row=i, column=1, value=wn or '')
        ws.cell(row=i, column=2, value=wa or '')
        if fid and member:
            ws.cell(row=i, column=3, value='ID found — matched member record')
            for ci, col in enumerate(['Member ID', 'Full Name', 'Status', 'District', 'Email'], 4):
                ws.cell(row=i, column=ci, value=member.get(col, ''))
        elif fid:
            ws.cell(row=i, column=3, value=f'ID {fid} found in name but NOT in member CSV')
            ws.cell(row=i, column=4, value=fid)
        else:
            ws.cell(row=i, column=3, value='No member ID; no name/email match')
        style_data_row(ws, i, len(WC_NM_COLS))
    set_widths(ws, WC_W)
    ws.freeze_panes = 'A2'


def write_csv_sheet(ws, items, fill):
    write_header(ws, CSV_COLS, fill)
    for i, m in enumerate(items, 2):
        for ci, col in enumerate(CSV_COLS, 1):
            ws.cell(row=i, column=ci, value=m.get(col, ''))
        style_data_row(ws, i, len(CSV_COLS))
    set_widths(ws, CSV_W)
    ws.freeze_panes = 'A2'


def build_workbook(confirmed_by_id, guessed_by_mid, wc_no_match,
                   id_to_member, all_members, matched_ids, output_path):

    confirmed = list(confirmed_by_id.values())
    guessed   = list(guessed_by_mid.values())
    csv_no_match = [m for m in all_members if m['Member ID'].strip() not in matched_ids]

    confirmed_active = [x for x in confirmed if is_active(x[2])]
    confirmed_review = [x for x in confirmed if not is_active(x[2])]
    guessed_active   = [x for x in guessed   if is_active(x[2])]
    guessed_review   = [x for x in guessed   if not is_active(x[2])]
    csv_active       = [m for m in csv_no_match if is_active(m)]
    csv_rest         = [m for m in csv_no_match if not is_active(m)]

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "1. Confirmed Active"
    write_confirmed_sheet(ws1, confirmed_active, make_fill(PALETTE['confirmed_active']))

    ws2 = wb.create_sheet("2. Confirmed Review")
    write_confirmed_sheet(ws2, confirmed_review, make_fill(PALETTE['confirmed_review']),
                          highlight=True)

    ws3 = wb.create_sheet("3. Guessed Active")
    write_guessed_sheet(ws3, guessed_active, make_fill(PALETTE['guessed_active']))

    ws4 = wb.create_sheet("4. Guessed Review")
    write_guessed_sheet(ws4, guessed_review, make_fill(PALETTE['guessed_review']),
                        highlight=True)

    ws5 = wb.create_sheet("5. WeChat No CSV Match")
    write_wc_nomatch_sheet(ws5, wc_no_match, id_to_member,
                           make_fill(PALETTE['wc_nomatch']))

    ws6 = wb.create_sheet("6. CSV No WeChat - Active")
    write_csv_sheet(ws6, csv_active, make_fill(PALETTE['csv_nomatch_active']))

    ws7 = wb.create_sheet("7. CSV No WeChat - Rest")
    write_csv_sheet(ws7, csv_rest, make_fill(PALETTE['csv_nomatch_rest']))

    wb.save(output_path)

    return {
        'confirmed_active': len(confirmed_active),
        'confirmed_review': len(confirmed_review),
        'guessed_active':   len(guessed_active),
        'guessed_review':   len(guessed_review),
        'wc_no_match':      len(wc_no_match),
        'csv_active':       len(csv_active),
        'csv_rest':         len(csv_rest),
    }


# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description='Match WeChat group members to member CSV database.'
    )
    parser.add_argument('--csv',     required=True,
                        help='Path to member database CSV')
    parser.add_argument('--wechat',  required=True,
                        help='Path to WeChat members CSV (columns: wechat_name, wechat_alias)')
    parser.add_argument('--output',  default='wechat_matching_output.xlsx',
                        help='Output Excel file path')
    parser.add_argument('--verbose', action='store_true',
                        help='Print all guessed matches for audit')
    args = parser.parse_args()

    print(f"Loading member CSV:    {args.csv}")
    id_to_member, all_members = load_csv_members(args.csv)
    print(f"  → {len(all_members)} unique member records")

    print(f"Loading WeChat list:   {args.wechat}")
    wechat_list = load_wechat_members(args.wechat)
    print(f"  → {len(wechat_list)} unique WeChat entries")

    print("Running matching...")
    confirmed_by_id, guessed_by_mid, wc_no_match = run_matching(
        wechat_list, id_to_member, all_members
    )
    matched_ids = set(confirmed_by_id.keys()) | set(guessed_by_mid.keys())

    if args.verbose:
        print("\n=== ALL GUESSED MATCHES (for human audit) ===")
        for wn, wa, m, conf, r in sorted(guessed_by_mid.values(), key=lambda x: -x[3]):
            print(f"  {conf:3}%  {wn!r:40}  →  {m['Member ID']} {m['Full Name']:28}  |  {r}")

    print(f"\nBuilding Excel: {args.output}")
    stats = build_workbook(confirmed_by_id, guessed_by_mid, wc_no_match,
                           id_to_member, all_members, matched_ids, args.output)

    print("\n=== RESULTS ===")
    print(f"  Tab 1  Confirmed Active:      {stats['confirmed_active']:4}")
    print(f"  Tab 2  Confirmed Review:       {stats['confirmed_review']:4}  ← members in group but inactive/expired!")
    print(f"  Tab 3  Guessed Active:         {stats['guessed_active']:4}  ← review for accuracy")
    print(f"  Tab 4  Guessed Review:         {stats['guessed_review']:4}  ← review for accuracy")
    print(f"  Tab 5  WeChat No CSV Match:    {stats['wc_no_match']:4}  ← not in member database")
    print(f"  Tab 6  CSV No WeChat Active:   {stats['csv_active']:4}  ← active members missing from group!")
    print(f"  Tab 7  CSV No WeChat Rest:     {stats['csv_rest']:4}")
    print(f"\nDone → {args.output}")


if __name__ == '__main__':
    main()
